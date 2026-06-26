from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BLOCKS = [
    (
        "postes",
        "Postes",
        1,
        3,
        {"ALTURA": "altura_m", "RESISTENCIA": "resistencia_dan", "TIPO": "tipo"},
    ),
    (
        "estruturas_mt",
        "Estruturas MT",
        5,
        9,
        {
            "NOME": "nome",
            "ESTILO DE REDE": "estilo_rede",
            "TIPO DE REDE": "tipo_rede",
            "CABOS": "cabos",
            "ANCORAGEM": "ancoragem",
        },
    ),
    (
        "estruturas_bt",
        "Estruturas BT",
        11,
        14,
        {
            "NOME": "nome",
            "TIPO DE REDE": "tipo_rede",
            "CABOS": "cabos",
            "ANCORAGEM": "ancoragem",
        },
    ),
    (
        "cabos",
        "Cabos",
        16,
        20,
        {
            "NOME": "nome",
            "TIPO DE REDE": "tipo_rede",
            "ESTILO DE REDE": "estilo_rede",
            "TENSÃO DA REDE": "tensao_rede",
            "FATOR DE CONDENAR": "fator_condenar",
        },
    ),
]


def build_definitions(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook.active
    entities: dict[str, Any] = {}

    for key, display_name, start_col, end_col, mapping in BLOCKS:
        headers = [sheet.cell(2, col).value for col in range(start_col, end_col + 1)]
        records: list[dict[str, Any]] = []
        for row_number in range(3, sheet.max_row + 1):
            values = [sheet.cell(row_number, col).value for col in range(start_col, end_col + 1)]
            if not any(value is not None for value in values):
                continue
            record = {
                mapping[header]: values[index]
                for index, header in enumerate(headers)
                if header in mapping and values[index] is not None
            }
            records.append(record)
        entities[key] = {"display_name": display_name, "records": records}

    return {"source": workbook_path.name, "entities": entities}


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa entidades da planilha de exemplos do ProjectHandler.")
    parser.add_argument("input", type=Path, help="Caminho para ARQUIVOS PARA PROJETO.xlsx")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("src/projecthandler/data/entity_definitions.json"),
        help="Arquivo JSON de saida.",
    )
    args = parser.parse_args()

    definitions = build_definitions(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(definitions, ensure_ascii=True, indent=2), encoding="utf-8")
    print(f"Arquivo gerado: {args.output}")


if __name__ == "__main__":
    main()
